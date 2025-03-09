import os
import time
import json
import argparse
import pandas as pd
from tqdm import tqdm
from datasets import load_dataset
from openai import OpenAI
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import re

def load_hf_dataset(dataset_id, split="train"):
    """Load a dataset from Hugging Face."""
    print(f"Loading dataset {dataset_id}...")
    dataset = load_dataset(dataset_id, split=split)
    return dataset

def setup_openai_client(api_key=None, base_url=None):
    """Set up the OpenAI client with the given API key and optional base URL."""
    # Set OpenAI's API key and API base to use vLLM's API server.
    openai_api_key = "EMPTY"
    openai_api_base = "http://localhost:8000/v1"
    
    client = OpenAI(
        api_key=openai_api_key,
        base_url=openai_api_base,
    )
    return client

def extract_qa_pairs_from_text(text):
    """Extract question-answer pairs from the text response."""
    qa_pairs = []
    
    # Try parsing as JSON first
    try:
        # Check if the text contains a JSON object
        json_match = re.search(r'({[\s\S]*})', text)
        if json_match:
            json_text = json_match.group(1)
            data = json.loads(json_text)
            
            if "qa_pairs" in data:
                return data["qa_pairs"]
            
            # Look for any array of objects with question and answer fields
            for key, value in data.items():
                if isinstance(value, list) and all(isinstance(item, dict) and "question" in item and "answer" in item for item in value):
                    return value
    except Exception as e:
        print(f"JSON parsing failed: {e}")
    
    # If JSON parsing fails, try regex pattern matching
    try:
        # Look for question-answer patterns in the text
        questions = re.findall(r'(?:^|\n)(?:Question\s*\d*:?\s*|Q\d*:?\s*|"\s*question\s*"\s*:\s*")(.*?)(?:(?=\n(?:Answer|A)\s*\d*:)|(?=")|$)', text, re.IGNORECASE)
        answers = re.findall(r'(?:^|\n)(?:Answer\s*\d*:?\s*|A\d*:?\s*|"\s*answer\s*"\s*:\s*")(.*?)(?:(?=\n(?:Question|Q)\s*\d*:)|(?=")|$)', text, re.IGNORECASE)
        
        # If we found matching questions and answers
        if questions and answers and len(questions) == len(answers):
            for q, a in zip(questions, answers):
                # Clean up the strings
                q = q.strip().strip('"').strip()
                a = a.strip().strip('"').strip()
                if q and a:
                    qa_pairs.append({"question": q, "answer": a})
            return qa_pairs
        
        # Try another pattern - numbered list
        pattern = r'(?:\d+\.|\(\d+\))\s+(.*?)\s*\n\s*(?:Answer:)?\s*([\s\S]*?)(?=(?:\d+\.|\(\d+\))|$)'
        matches = re.findall(pattern, text)
        
        if matches:
            for q, a in matches:
                q = q.strip()
                a = a.strip()
                if q and a:
                    qa_pairs.append({"question": q, "answer": a})
            return qa_pairs
            
    except Exception as e:
        print(f"Regex parsing failed: {e}")
    
    # If all automated parsing fails, just create a single QA pair with the raw text
    if not qa_pairs:
        print("Could not parse QA pairs from text, saving raw response")
        qa_pairs.append({
            "question": "Raw response from model:",
            "answer": text
        })
    
    return qa_pairs

def generate_qa_pairs(client, text, n_questions=25, model="gpt-3.5-turbo", max_retries=3, retry_delay=5):
    """Generate question-answer pairs from the given text using the OpenAI API."""
    prompt = f"""
    You are an expert in Pakistani law, particularly the Constitution of Pakistan and the Pakistan Penal Code.
    
    Given the following text about Pakistani law, generate {n_questions} specific, detailed questions with accurate answers that could be used to test 
    someone's knowledge of Pakistan's constitution and penal code.
    
    The questions should:
    1. Be diverse and cover different aspects of the text
    2. Range from factual to analytical or interpretive
    3. Use precise legal terminology where appropriate
    4. Be clearly answerable based on the text provided
    5. Focus specifically on constitutional provisions or penal code sections when possible
    6. Do not add references and cross-references in your answer. Provide Direct Answers
    7. Provide direct and concise answer to each question
    
    The answers should:
    1. Be comprehensive and accurate
    2. Cite specific sections of law where relevant
    3. Explain legal principles clearly
    4. Be factually correct according to Pakistani law
    5. Provide direct and concise answer.
    
    TEXT:
    {text}
    
    FORMAT YOUR RESPONSE AS A JSON ARRAY OF OBJECTS, each with "question" and "answer" fields.
    Example format:
    {{
      "qa_pairs": [
        {{"question": "What is Article 25 of the Constitution of Pakistan?", "answer": "Article 25 guarantees equality before law and equal protection of law to all citizens."}},
        {{"question": "What penalty does Section 302 of Pakistan Penal Code prescribe?", "answer": "Section 302 of Pakistan Penal Code prescribes death or imprisonment for life as punishment for murder."}}
      ]
    }}
    """
    
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model="lambdalabs/Llama-3.3-70B-Instruct-AWQ-4bit",
                messages=[
                    {"role": "system", "content": "You are a legal expert assistant specializing in Pakistani law."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
            )
            
            # Parse the response content
            response_text = response.choices[0].message.content
            
            # Try to extract QA pairs from the text
            qa_pairs = extract_qa_pairs_from_text(response_text)
            
            # Save the raw response to a text file for debugging
            with open(f"response_{time.strftime('%Y%m%d_%H%M%S')}.txt", "w", encoding="utf-8") as f:
                f.write(response_text)
                
            return qa_pairs
            
        except Exception as e:
            print(f"Attempt {attempt+1}/{max_retries} failed: {e}")
            if attempt < max_retries - 1:
                print(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
            else:
                print("Max retries reached. Returning empty list.")
                return []

def setup_excel_file(excel_file):
    """Set up the Excel file for saving results."""
    if os.path.exists(excel_file):
        # Load existing workbook
        wb = load_workbook(excel_file)
        if 'QA Pairs' in wb.sheetnames:
            ws = wb['QA Pairs']
            # Get the next available row
            next_row = ws.max_row + 1
        else:
            # Create a new sheet
            ws = wb.create_sheet('QA Pairs')
            ws.append(['Question', 'Answer'])
            next_row = 2
    else:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'QA Pairs'
        ws.append(['Question', 'Answer'])
        next_row = 2
    
    wb.save(excel_file)
    return next_row

def save_qa_pairs_to_excel(qa_pairs, excel_file, start_row=2):
    """Save the given QA pairs to an Excel file."""
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb['QA Pairs']
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'QA Pairs'
        ws.append(['Question', 'Answer'])
    
    current_row = start_row
    
    for pair in qa_pairs:
        question = pair.get('question', '')
        answer = pair.get('answer', '')
        ws.append([question, answer])
        current_row += 1
    
    wb.save(excel_file)
    return current_row

def process_dataset(dataset, client, json_output_file, excel_output_file, model, target_questions=2000, start_idx=0, end_idx=None, chunk_size=1):
    """Process the dataset and generate QA pairs for each row."""
    results = []
    total_qa_pairs = 0
    
    # Convert to pandas DataFrame for easier handling
    df = pd.DataFrame(dataset)
    
    if end_idx is None:
        end_idx = len(df)
    else:
        end_idx = min(end_idx, len(df))
    
    # Subset the dataframe
    df = df.iloc[start_idx:end_idx]
    
    print(f"Processing {len(df)} rows from index {start_idx} to {end_idx-1}...")
    
    # Set up the Excel file and get the starting row
    excel_row = setup_excel_file(excel_output_file)
    
    # Process in chunks and save intermediate results
    for i in tqdm(range(0, len(df), chunk_size)):
        chunk_df = df.iloc[i:i+chunk_size]
        chunk_qa_pairs = []
        
        for _, row in chunk_df.iterrows():
            # Check if we've reached the target number of questions
            if total_qa_pairs >= target_questions:
                print(f"Reached target of {target_questions} questions. Stopping.")
                break
                
            chunk_id = row['chunk_id'] if 'chunk_id' in row else 'unknown'
            text = row['text']
            
            # Skip if text is too short
            if len(text.split()) < 50:
                print(f"Skipping chunk {chunk_id} as it's too short ({len(text.split())} words)")
                continue
            
            # Calculate how many questions to generate from this chunk
            # Adjust the number to reach the target total
            remaining = target_questions - total_qa_pairs
            n_questions = min(25, remaining)  # Generate up to 25 questions per chunk
            
            if n_questions <= 0:
                break
                
            # Generate QA pairs
            qa_pairs = generate_qa_pairs(client, text, n_questions=n_questions, model=model)
            
            # Save to Excel on the fly
            if qa_pairs:
                excel_row = save_qa_pairs_to_excel(qa_pairs, excel_output_file, excel_row)
                
                result = {
                    'chunk_id': chunk_id,
                    'original_text': text,
                    'qa_pairs': qa_pairs
                }
                
                results.append(result)
                chunk_qa_pairs.extend(qa_pairs)
                total_qa_pairs += len(qa_pairs)
                
                print(f"Generated {len(qa_pairs)} QA pairs from chunk {chunk_id}. Total: {total_qa_pairs}/{target_questions}")
            else:
                print(f"Failed to generate QA pairs from chunk {chunk_id}.")
            
            if total_qa_pairs >= target_questions:
                print(f"Reached target of {target_questions} questions. Stopping.")
                break
        
        # Save intermediate JSON results
        intermediate_file = f"{json_output_file.rsplit('.', 1)[0]}_intermediate_{start_idx+i}.json"
        with open(intermediate_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"Saved intermediate results to {intermediate_file}")
        
        if total_qa_pairs >= target_questions:
            break
    
    # Save final JSON results
    with open(json_output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"Saved JSON results to {json_output_file}")
    print(f"Saved Excel results to {excel_output_file}")
    print(f"Total QA pairs generated: {total_qa_pairs}")
    
    return results

def main():
    parser = argparse.ArgumentParser(description='Generate question-answer pairs from a Hugging Face dataset')
    parser.add_argument('--dataset', type=str, default='Syed-Hasan-8503/FYP-PPC-pretrain-corpus',
                        help='Hugging Face dataset ID')
    parser.add_argument('--split', type=str, default='train',
                        help='Dataset split to use')
    parser.add_argument('--json-output', type=str, default='pakistan_legal_qa.json',
                        help='Output JSON file path')
    parser.add_argument('--excel-output', type=str, default='pakistan_legal_qa.xlsx',
                        help='Output Excel file path')
    parser.add_argument('--target-questions', type=int, default=2000,
                        help='Target number of questions to generate')
    parser.add_argument('--start-idx', type=int, default=0,
                        help='Starting index in the dataset')
    parser.add_argument('--end-idx', type=int, default=None,
                        help='Ending index in the dataset (exclusive)')
    parser.add_argument('--chunk-size', type=int, default=1,
                        help='Number of rows to process before saving intermediate results')
    
    args = parser.parse_args()
    
    # Load dataset
    dataset = load_hf_dataset(args.dataset, args.split)
    
    # Setup client
    client = setup_openai_client()
    
    # Process dataset
    process_dataset(
        dataset=dataset,
        client=client,
        json_output_file=args.json_output,
        excel_output_file=args.excel_output,
        model="lambdalabs/Llama-3.3-70B-Instruct-AWQ-4bit",
        target_questions=args.target_questions,
        start_idx=args.start_idx,
        end_idx=args.end_idx,
        chunk_size=args.chunk_size
    )

if __name__ == "__main__":
    main()
    
    # Process dataset
    process_dataset(
        dataset=dataset,
        client=client,
        json_output_file=args.json_output,
        excel_output_file=args.excel_output,
        model=args.model,
        target_questions=args.target_questions,
        start_idx=args.start_idx,
        end_idx=args.end_idx,
        chunk_size=args.chunk_size
    )

if __name__ == "__main__":
    main()